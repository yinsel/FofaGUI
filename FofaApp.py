# -*- coding: utf-8 -*-
import asyncio
import base64
import json
import os.path
import time
import aiohttp
import openpyxl
import rsa
from openpyxl.styles import Alignment
import openpyxl.utils
import argparse

# 手动输入查询语法
QUERY = ''

# 全局变量
PRIVATE_STR = "-----BEGIN RSA PRIVATE KEY-----\r\nMIIEogIBAAKCAQEAv0xjefuBTF6Ox940ZqLLUFFBDtTcB9dAfDjWgyZ2A55K+VdG\r\nc1L5LqJWuyRkhYGFTlI4K5hRiExvjXuwIEed1norp5cKdeTLJwmvPyFgaEh7Ow19\r\nTu9sTR5hHxThjT8ieArB2kNAdp8Xoo7O8KihmBmtbJ1umRv2XxG+mm2ByPZFlTdW\r\nRFU38oCPkGKlrl/RzOJKRYMv10s1MWBPY6oYkRiOX/EsAUVae6zKRqNR2Q4HzJV8\r\ngOYMPvqkau8hwN8i6r0z0jkDGCRJSW9djWk3Byi3R2oSdZ0IoS+91MFtKvWYdnNH\r\n2Ubhlnu1P+wbeuIFdp2u7ZQOtgPX0mtQ263e5QIDAQABAoIBAD67GwfeTMkxXNr3\r\n5/EcQ1XEP3RQoxLDKHdT4CxDyYFoQCfB0e1xcRs0ywI1be1FyuQjHB5Xpazve8lG\r\nnTwIoB68E2KyqhB9BY14pIosNMQduKNlygi/hKFJbAnYPBqocHIy/NzJHvOHOiXp\r\ndL0AX3VUPkWW3rTAsar9U6aqcFvorMJQ2NPjijcXA0p1MlZAZKODO2wqidfQ487h\r\nxy0ZkriYVi419j83a1cCK0QocXiUUeQM6zRNgQv7LCmrFo2X4JEzlujEveqvsDC4\r\nMBRgkK2lNH+AFuRwOEr4PIlk9rrpHA4O1V13P3hJpH5gxs5oLLM1CWWG9YWLL44G\r\nzD9Tm8ECgYEA8NStMXyAmHLYmd2h0u5jpNGbegf96z9s/RnCVbNHmIqh/pbXizcv\r\nmMeLR7a0BLs9eiCpjNf9hob/JCJTms6SmqJ5NyRMJtZghF6YJuCSO1MTxkI/6RUw\r\nmrygQTiF8RyVUlEoNJyhZCVWqCYjctAisEDaBRnUTpNn0mLvEXgf1pUCgYEAy1kE\r\nd0YqGh/z4c/D09crQMrR/lvTOD+LRMf9lH+SkScT0GzdNIT5yuscRwKsnE6SpC5G\r\nySJFVhCnCBsQqq+ohsrXt8a99G7ePTMSAGK3QtC7QS3liDmvPBk6mJiLrKiRAZos\r\nvgPg7nTP8VuF0ZIKzkdWbGoMyNxVFZXovQ8BYxECgYBvCR9xGX4Qy6KiDlV18wNu\r\nElYkxVqFBBE0AJRg/u+bnQ9jWhi2zxLa1eWZgtss80c876I8lbkGNWedOVZioatm\r\nMFLC4bFalqyZWyO7iP7i60LKvfDJfkOSlDUu3OikahFOiqyG1VBz4+M4U500alIU\r\nAVKD14zTTZMopQSkgUXsoQKBgHd8RgiD3Qde0SJVv97BZzP6OWw5rqI1jHMNBK72\r\nSzwpdxYYcd6DaHfYsNP0+VIbRUVdv9A95/oLbOpxZNi2wNL7a8gb6tAvOT1Cvggl\r\n+UM0fWNuQZpLMvGgbXLu59u7bQFBA5tfkhLr5qgOvFIJe3n8JwcrRXndJc26OXil\r\n0Y3RAoGAJOqYN2CD4vOs6CHdnQvyn7ICc41ila/H49fjsiJ70RUD1aD8nYuosOnj\r\nwbG6+eWekyLZ1RVEw3eRF+aMOEFNaK6xKjXGMhuWj3A9xVw9Fauv8a2KBU42Vmcd\r\nt4HRyaBPCQQsIoErdChZj8g7DdxWheuiKoN4gbfK4W1APCcuhUA=\r\n-----END RSA PRIVATE KEY-----"
APP_ID = "9e9fb94330d97833acfbc041ee1a76793f1bc691"
page_size = 10
params_temp = {
    "full": "false",
    "page": "",
    "size": "10",
    "q": "",
    "qbase64": "",
    "t": ""
}
table_data = [["标题", "域名/IP:端口", "IP", "指纹", "国家"]]

# 获取请求参数
def get_params(page, t, q):
    params_copy = params_temp.copy()
    params_copy['page'] = str(page)
    params_copy['t'] = t
    if "&&" in q:
        p = q.split("&&")
        for index, e in enumerate(p):
            e = e.strip()
            if index == 0:
                params_copy["q"] = e
                params_copy['qbase64'] = base64.b64encode(q.encode()).decode()
            else:
                if "||" in e:
                    s = e.split("||")
                    for o in s:
                        o = o.strip()
                        if "=" in o:
                            k = o.split("=")[0]
                            v = o.split("=")[1]
                            params_copy[k] = v
                        else:
                            params_copy[o] = ""
                else:
                    if "=" in e:
                        k = e.split("=")[0]
                        v = e.split("=")[1]
                        params_copy[k] = v
                    else:
                        params_copy[e] = ""
    else:
        params_copy["q"] = q
        params_copy['qbase64'] = base64.b64encode(q.encode()).decode()
    sorted_params = sorted(params_copy.items(), key=lambda x: x[0][0])
    for_sign = ""
    for s in sorted_params:
        for_sign += s[0] + s[1]
    private_key = rsa.PrivateKey.load_pkcs1(PRIVATE_STR.encode())
    sign = base64.b64encode(
        rsa.sign_hash(rsa.compute_hash(for_sign.encode(), "SHA-256"), private_key, "SHA-256")).decode()
    params_copy['app_id'] = APP_ID
    params_copy['sign'] = sign
    return params_copy


# 文件校验
def check():
    if not os.path.exists("cookie/fofa.json"):
        print("fofa.json文件不存在!")
        return False
    else:
        with open("cookie/fofa.json", "r", encoding="utf-8") as file:
            if len(file.read()) == 0:
                print("请在fofa.json中填入cookie")
                return False
    return True

# 获取cookie
def get_cookies():
    cookies = {}
    with open("cookie/fofa.json", "r", encoding="utf-8") as file:
        json_data = json.load(file)
        for cookie in json_data:
            cookies[cookie['name']] = cookie['value']
        return cookies

# 请求
async def request(client, sec, page, cookies, t, q):
    params = get_params(page, t, q)
    await asyncio.sleep(sec)
    resp = await client.get('https://api.fofa.info/v1/search', cookies=cookies, params=params,
                            headers={"Authorization": cookies['fofa_token']})
    result = await resp.json()
    try:
        data_list = result['data']['assets']
    except TypeError:
        return None
    return page, data_list

# excel处理
def output_excel(data, t, q):
    table = openpyxl.Workbook()
    fofa_data = table.active
    row = len(data)

    # 添加标题到第一行
    title = f"查询内容：{q}"
    fofa_data.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)  # 合并单元格
    title_cell = fofa_data.cell(row=1, column=1, value=title)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    for i in range(1, row + 2):  # 增加 1 以留出标题的空间
        fofa_data.row_dimensions[i].height = 15

    for i in range(1, 5 + 1):
        # 获取列的字母标识（A、B、C...）
        column_letter = openpyxl.utils.get_column_letter(i)
        # 设置列的宽度
        fofa_data.column_dimensions[column_letter].width = 40

    for col in fofa_data.columns:
        for cell in col:
            # 设置单元格内容居中对齐
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, row_data in enumerate(data, start=2):  # 从第二行开始添加数据
        for col_idx, value in enumerate(row_data, start=1):
            fofa_data.cell(row=row_idx, column=col_idx, value=value)

    if not os.path.exists("查询结果"):
        os.makedirs("查询结果")
    table.save(f"查询结果/{t}.xlsx")

# 主函数
async def main():
    # 命令行解析参数
    argparser = argparse.ArgumentParser(description="Fofa查询工具 by yinsel",formatter_class=argparse.RawTextHelpFormatter)
    argparser.add_argument("-q",dest="query",type=str,help="查询语法",required=False)
    args = argparser.parse_args()
    if not any(vars(args).values()) and QUERY == "":
        argparser.print_help()
        return
    query = args.query
    if QUERY != "":
        query = QUERY

    # 判断cookie是否存在
    if not check():
        return
    
    # 创建请求
    async with aiohttp.ClientSession() as session:
        cookies = get_cookies()
        t = str(int(time.time() * 1000))
        task_list = []
        sec = 0
        for i in range(1, 7):
            sec += 0.5
            req = request(session, sec, i, cookies, t, query)
            task = asyncio.create_task(req)
            task_list.append(task)
        try:
            print("查询中...")
            data_lists = await asyncio.gather(*task_list)
            data_lists.sort()
        except TypeError:
            print("搜索结果为空或Cookie过期")
            return

    # 结果处理
    for s in data_lists:
        data_list = s[1]
        for e in data_list:
            title = e.get('title', "")
            id = e.get('id', "")
            ip = e.get("ip", "")
            country = e.get("country")
            if e["app_servers"]:
                servers = ' '.join(str(server.get('name', "")) for server in e["app_servers"])
            else:
                servers = ""
            table_data.append([title, id, ip, servers, country])
        output_excel(table_data, t, query)
    print(f"查询成功!\n查询内容: {query}\n文件路径为: 查询结果/{t}.xlsx")
    os.startfile(f"查询结果\\{t}.xlsx")

# 入口
asyncio.run(main())