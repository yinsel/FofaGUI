# -*- coding: utf-8 -*-
import asyncio
import base64
import functools
import json
import os.path
import sys
import time
import aiohttp
import openpyxl
import rsa
from PyQt5 import QtCore, QtWidgets
import qasync
from PyQt5.QtWidgets import QMessageBox
from openpyxl.styles import Alignment
from qasync import asyncSlot, asyncClose
import openpyxl.utils
import resource_rc
# Form implementation generated from reading ui file 'FofaGUIApp.ui'
#
# Created by: PyQt5 UI code generator 5.15.9
#
# WARNING: Any manual changes made to this file will be lost when pyuic5 is
# run again.  Do not edit this file unless you know what you are doing.

# 全局变量
PRIVATE_STR = "-----BEGIN RSA PRIVATE KEY-----\r\nMIIEogIBAAKCAQEAv0xjefuBTF6Ox940ZqLLUFFBDtTcB9dAfDjWgyZ2A55K+VdG\r\nc1L5LqJWuyRkhYGFTlI4K5hRiExvjXuwIEed1norp5cKdeTLJwmvPyFgaEh7Ow19\r\nTu9sTR5hHxThjT8ieArB2kNAdp8Xoo7O8KihmBmtbJ1umRv2XxG+mm2ByPZFlTdW\r\nRFU38oCPkGKlrl/RzOJKRYMv10s1MWBPY6oYkRiOX/EsAUVae6zKRqNR2Q4HzJV8\r\ngOYMPvqkau8hwN8i6r0z0jkDGCRJSW9djWk3Byi3R2oSdZ0IoS+91MFtKvWYdnNH\r\n2Ubhlnu1P+wbeuIFdp2u7ZQOtgPX0mtQ263e5QIDAQABAoIBAD67GwfeTMkxXNr3\r\n5/EcQ1XEP3RQoxLDKHdT4CxDyYFoQCfB0e1xcRs0ywI1be1FyuQjHB5Xpazve8lG\r\nnTwIoB68E2KyqhB9BY14pIosNMQduKNlygi/hKFJbAnYPBqocHIy/NzJHvOHOiXp\r\ndL0AX3VUPkWW3rTAsar9U6aqcFvorMJQ2NPjijcXA0p1MlZAZKODO2wqidfQ487h\r\nxy0ZkriYVi419j83a1cCK0QocXiUUeQM6zRNgQv7LCmrFo2X4JEzlujEveqvsDC4\r\nMBRgkK2lNH+AFuRwOEr4PIlk9rrpHA4O1V13P3hJpH5gxs5oLLM1CWWG9YWLL44G\r\nzD9Tm8ECgYEA8NStMXyAmHLYmd2h0u5jpNGbegf96z9s/RnCVbNHmIqh/pbXizcv\r\nmMeLR7a0BLs9eiCpjNf9hob/JCJTms6SmqJ5NyRMJtZghF6YJuCSO1MTxkI/6RUw\r\nmrygQTiF8RyVUlEoNJyhZCVWqCYjctAisEDaBRnUTpNn0mLvEXgf1pUCgYEAy1kE\r\nd0YqGh/z4c/D09crQMrR/lvTOD+LRMf9lH+SkScT0GzdNIT5yuscRwKsnE6SpC5G\r\nySJFVhCnCBsQqq+ohsrXt8a99G7ePTMSAGK3QtC7QS3liDmvPBk6mJiLrKiRAZos\r\nvgPg7nTP8VuF0ZIKzkdWbGoMyNxVFZXovQ8BYxECgYBvCR9xGX4Qy6KiDlV18wNu\r\nElYkxVqFBBE0AJRg/u+bnQ9jWhi2zxLa1eWZgtss80c876I8lbkGNWedOVZioatm\r\nMFLC4bFalqyZWyO7iP7i60LKvfDJfkOSlDUu3OikahFOiqyG1VBz4+M4U500alIU\r\nAVKD14zTTZMopQSkgUXsoQKBgHd8RgiD3Qde0SJVv97BZzP6OWw5rqI1jHMNBK72\r\nSzwpdxYYcd6DaHfYsNP0+VIbRUVdv9A95/oLbOpxZNi2wNL7a8gb6tAvOT1Cvggl\r\n+UM0fWNuQZpLMvGgbXLu59u7bQFBA5tfkhLr5qgOvFIJe3n8JwcrRXndJc26OXil\r\n0Y3RAoGAJOqYN2CD4vOs6CHdnQvyn7ICc41ila/H49fjsiJ70RUD1aD8nYuosOnj\r\nwbG6+eWekyLZ1RVEw3eRF+aMOEFNaK6xKjXGMhuWj3A9xVw9Fauv8a2KBU42Vmcd\r\nt4HRyaBPCQQsIoErdChZj8g7DdxWheuiKoN4gbfK4W1APCcuhUA=\r\n-----END RSA PRIVATE KEY-----"
APP_ID = "9e9fb94330d97833acfbc041ee1a76793f1bc691"
page_size = 10
params_temp = {
    "full": "false",
    "page": "",
    "size": "10",
    "q": "",
    "qbase64": "",
    "t": ""
}
table_data = [["标题", "域名/IP:端口", "IP", "指纹", "国家"]]


def get_params(page, t, q):
    params_copy = params_temp.copy()
    params_copy['page'] = str(page)
    params_copy['t'] = t
    if "&&" in q:
        p = q.split("&&")
        for index, e in enumerate(p):
            e = e.strip()
            if index == 0:
                params_copy["q"] = e
                params_copy['qbase64'] = base64.b64encode(q.encode()).decode()
            else:
                if "||" in e:
                    s = e.split("||")
                    for o in s:
                        o = o.strip()
                        if "=" in o:
                            k = o.split("=")[0]
                            v = o.split("=")[1]
                            params_copy[k] = v
                        else:
                            params_copy[o] = ""
                else:
                    if "=" in e:
                        k = e.split("=")[0]
                        v = e.split("=")[1]
                        params_copy[k] = v
                    else:
                        params_copy[e] = ""
    else:
        params_copy["q"] = q
        params_copy['qbase64'] = base64.b64encode(q.encode()).decode()
    print(params_copy)
    sorted_params = sorted(params_copy.items(), key=lambda x: x[0][0])
    for_sign = ""
    for s in sorted_params:
        for_sign += s[0] + s[1]
    private_key = rsa.PrivateKey.load_pkcs1(PRIVATE_STR.encode())
    sign = base64.b64encode(
        rsa.sign_hash(rsa.compute_hash(for_sign.encode(), "SHA-256"), private_key, "SHA-256")).decode()
    params_copy['app_id'] = APP_ID
    params_copy['sign'] = sign
    return params_copy

def show_message(type_, message):
    message_box = QMessageBox()
    message_box.setWindowTitle('提示')
    message_box.setText(message)
    message_box.setIcon(type_)
    message_box.exec_()


def check():
    if not os.path.exists("cookie/fofa.json"):
        show_message(type_=QMessageBox.Warning, message="fofa.json文件不存在!")
        return False
    else:
        with open("cookie/fofa.json", "r", encoding="utf-8") as file:
            if len(file.read()) == 0:
                show_message(type_=QMessageBox.Warning, message="请在fofa.json中填入cookie")
                return False
    return True


def get_cookies():
    cookies = {}
    with open("cookie/fofa.json", "r", encoding="utf-8") as file:
        json_data = json.load(file)
        for cookie in json_data:
            cookies[cookie['name']] = cookie['value']
        return cookies


async def request(client, sec, page, cookies, t, q):
    params = get_params(page, t, q)
    await asyncio.sleep(sec)
    resp = await client.get('https://api.fofa.info/v1/search', cookies=cookies, params=params,
                            headers={"Authorization": cookies['fofa_token']})
    result = await resp.json()
    try:
        data_list = result['data']['assets']
    except TypeError:
        return None
    return page, data_list


def output_excel(data, t, q):
    table = openpyxl.Workbook()
    fofa_data = table.active
    row = len(data)

    # 添加标题到第一行
    title = f"查询内容：{q}"
    fofa_data.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)  # 合并单元格
    title_cell = fofa_data.cell(row=1, column=1, value=title)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    for i in range(1, row + 2):  # 增加 1 以留出标题的空间
        fofa_data.row_dimensions[i].height = 15

    for i in range(1, 5 + 1):
        # 获取列的字母标识（A、B、C...）
        column_letter = openpyxl.utils.get_column_letter(i)
        # 设置列的宽度
        fofa_data.column_dimensions[column_letter].width = 40

    for col in fofa_data.columns:
        for cell in col:
            # 设置单元格内容居中对齐
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, row_data in enumerate(data, start=2):  # 从第二行开始添加数据
        for col_idx, value in enumerate(row_data, start=1):
            fofa_data.cell(row=row_idx, column=col_idx, value=value)

    if not os.path.exists("查询结果"):
        os.makedirs("查询结果")
    table.save(f"查询结果/{t}.xlsx")


class UiWindow(object):
    def setupUi(self, Window):
        from PyQt5.QtGui import QIcon
        icon = QIcon(":/fofa.ico")
        Window.setWindowIcon(icon)
        Window.setObjectName("Window")
        Window.setWindowModality(QtCore.Qt.ApplicationModal)
        Window.resize(680, 110)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(Window.sizePolicy().hasHeightForWidth())
        Window.setSizePolicy(sizePolicy)
        Window.setMinimumSize(QtCore.QSize(680, 110))
        Window.setMaximumSize(QtCore.QSize(680, 110))
        self.search_edit = QtWidgets.QLineEdit(Window)
        self.search_edit.setGeometry(QtCore.QRect(100, 40, 451, 31))
        self.search_edit.setObjectName("search_edit")
        self.search_label = QtWidgets.QLabel(Window)
        self.search_label.setGeometry(QtCore.QRect(20, 40, 71, 31))
        self.search_label.setObjectName("search_label")
        self.search_button = QtWidgets.QPushButton(Window)
        self.search_button.setGeometry(QtCore.QRect(570, 40, 93, 31))
        self.search_button.setObjectName("search_button")


        self.retranslateUi(Window)
        QtCore.QMetaObject.connectSlotsByName(Window)

    def retranslateUi(self, Window):
        _translate = QtCore.QCoreApplication.translate
        Window.setWindowTitle(_translate("Window", "Fofa查询工具 by yinsel"))
        self.search_label.setText(_translate("Window", "查询内容："))
        self.search_button.setText(_translate("Window", "查询"))


class FofaGUIApp(QtWidgets.QMainWindow, UiWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.session = aiohttp.ClientSession(
            loop=asyncio.get_event_loop()
        )
        self.search_button.clicked.connect(self.search)

    @asyncClose
    async def closeEvent(self, event):
        await self.session.close()

    @asyncSlot()
    async def search(self):
        self.search_edit.setEnabled(False)
        self.search_button.setEnabled(False)
        self.search_button.setText("查询中...")
        if check():
            cookies = get_cookies()
            q = self.search_edit.text()
            t = str(int(time.time() * 1000))
            task_list = []
            sec = 0
            for i in range(1, 7):
                sec += 0.5
                req = request(self.session, sec, i, cookies, t, q)
                task = asyncio.create_task(req)
                task_list.append(task)
            try:
                data_lists = await asyncio.gather(*task_list)
                data_lists.sort()
            except TypeError:
                show_message(type_=QMessageBox.Warning, message="搜索结果为空或Cookie过期")
                self.search_button.setText("查询")
                self.search_button.setEnabled(True)
                self.search_edit.setEnabled(True)
                return
            for s in data_lists:
                data_list = s[1]
                for e in data_list:
                    title = e.get('title', "")
                    id = e.get('id', "")
                    ip = e.get("ip", "")
                    country = e.get("country")
                    if e["app_servers"]:
                        servers = ' '.join(str(server.get('name', "")) for server in e["app_servers"])
                    else:
                        servers = ""
                    table_data.append([title, id, ip, servers, country])
            output_excel(table_data, t, q)
            show_message(type_=QMessageBox.Information, message=f"查询成功!\n查询内容: {q}\n文件路径为: 查询结果/{t}.xlsx")
        self.search_button.setText("查询")
        self.search_button.setEnabled(True)
        self.search_edit.setEnabled(True)

async def main():
    def close_future(future, loop):
        loop.call_later(10, future.cancel)
        future.cancel()

    loop = asyncio.get_event_loop()
    future = asyncio.Future()

    app = qasync.QApplication.instance()
    if hasattr(app, "aboutToQuit"):
        getattr(app, "aboutToQuit").connect(
            functools.partial(close_future, future, loop)
        )

    window = FofaGUIApp()
    window.show()
    await future


if __name__ == '__main__':
    try:
        qasync.run(main())
    except asyncio.exceptions.CancelledError:
        sys.exit(0)
